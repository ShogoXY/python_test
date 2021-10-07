# dobra to teraz testujemy
import openpyxl
import os
import msvcrt
os.chdir("C:\\Users\\Dariusz\\Desktop\\Śmieci\\python")

# Path
wb = openpyxl.load_workbook('test1.xlsx')

# active worksheet data
ws = wb.active
print("podaj komentarz do wpisania")
komentarz = input()


#serial = (input("podaj wartość którą chcesz wyszukać\n"))
#print("by wyjść nie podawaj nic i wciśnij ENTER\n")

#    serial = input("podaj wartość którą chcesz wyszukać\n")
while True:
    def wordfinder(searchString):
        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                if searchString == ws.cell(i, j).value:
                    print(ws.cell(i, j))
                    test = ws.cell(i, 6)
                    test.value = komentarz
    print("podaj wartość którą chcesz wyszukać")
    wordfinder(input())
    wb.save('test1.xlsx')
else:
    break
