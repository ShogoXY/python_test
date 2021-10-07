# dobra to teraz testujemy
import openpyxl
import os
os.chdir("/home/fedora/python_test")

# Path
wb = openpyxl.load_workbook('test1.xlsx')

# active worksheet data
ws = wb.active
print("podaj komentarz do wpisania")
komentarz = input()


def wordfinder(searchString):
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if searchString == ws.cell(i, j).value:
                print(ws.cell(i, j))
                test = ws.cell(i, 6)
                test.value = komentarz


print("podaj wartość którą chcesz wyszukać")
wordfinder(input())
wb.save('test1.xlsx')
