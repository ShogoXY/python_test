# program do tworzenia karty gwarancyjnej w PDF na podstawie danych z Excel
import os
# import pandas as pd
# import docxptl
# import docx
import openpyxl
import sys

os.chdir("/home/fedora/git/python_test")
wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Arkusz1']
searchString = input("podaj wartość którą chcesz wyszukać\n")
if searchString not in wb.sheetnames:
    wb.create_sheet(searchString)
    ws_1 = wb[searchString]
    for i in range(1, ws.max_row + 1):
        if searchString == ws.cell(i, 5).value:
            commet = ws.cell(i, 11).value
            for i in range(1, ws.max_row + 1):
                if commet == ws.cell(i, 11).value:
                    pn = ws.cell(i, 4).value
                    sn = ws.cell(i, 5).value
                    gwara = ("Produkt --  " + pn +
                             " \nSeryjny --  " + sn + "\n")
                    print(gwara)
                    ws_1.cell(i, 1).value = pn
                    ws_1.cell(i, 2).value = sn

    count = 0
    for row in ws_1:
        if all([cell.value == None for cell in row]):
            count += 1
    ws_1.delete_rows(1, count)

else:
    for i in range(1, ws.max_row + 1):
        if searchString == ws.cell(i, 5).value:
            commet = ws.cell(i, 11).value
            for i in range(1, ws.max_row + 1):
                if commet == ws.cell(i, 11).value:
                    pn = ws.cell(i, 4).value
                    sn = ws.cell(i, 5).value
                    gwara = ("Produkt --  " + pn +
                             " \nSeryjny --  " + sn + "\n")
                    print(gwara)


wb.save('test.xlsx')
